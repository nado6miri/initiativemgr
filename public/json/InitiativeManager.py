import sys
import os
import copy
import time

#for Jira control
from jira import JIRA
from jira.exceptions import JIRAError

#pip install XlsxWriter

#xlsxwriter
#https://xlsxwriter.readthedocs.io/index.html

#openpyxl
#http://www.hanul93.com/openpyxl-basic/
#https://openpyxl.readthedocs.io/en/stable/defined_names.html



#http://hlm.lge.com/issue/rest/api/2/issue/GSWDIM-22476/
#http://hlm.lge.com/issue/rest/api/2/issue/TVPLAT-3963/
#http://hlm.lge.com/issue/rest/api/2/issue/TVPLAT-3963/editmeta
#http://hlm.lge.com/issue/rest/api/2/issue/TVPLAT-3963/?expand=changelog
#http://hlm.lge.com/qi/rest/api/2/issue/QEVENTSEVT-7232/ - Q

import openpyxl as xlsrd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color

from datetime import datetime
import copy

myfont = Font(name='맑은 고딕', size = 10, bold=True)
myalignment = Alignment(horizontal='center', vertical='center')
#myfill = PatternFill(patternType='solid', fgColor=Color('FFC000'))
myfill = PatternFill(patternType='solid', fgColor=Color('000000'))
myborder = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Initiative_webOS4.5_Initial_Dev_EpicList (42101)
#project = TVPLAT AND issuetype = Initiative AND fixVersion = "webOS TV 4.5 Initial" AND status not in (deferred, closed) AND (Component not in ("Deployment(manage)", _Governing) OR component is EMPTY)
# Initiative_webOS4.5_Initial_Dev_EpicList (42317)
# 관리 Epic 제외 : issuetype = Epic AND (component not in ("Deployment(manage)", _Governing) OR Component is EMPTY) AND issueFunction in linkedIssuesOf("project = TVPLAT and issuetype = Initiative and status not in (deferred, closed) and (component  not in (_Governing, 'Deployment(manage)') or Component is Empty) and fixVersion = 'webOS TV 4.5 Initial'")
# 관리 Epic 포함 : issuetype = Epic AND issueFunction in linkedIssuesOf("project = TVPLAT and issuetype = Initiative and status not in (deferred, closed) and (component  not in (_Governing, 'Deployment(manage)') or Component is Empty) and fixVersion = 'webOS TV 4.5 Initial'")
DevTracker = 'http://hlm.lge.com/issue'
QTracker = 'http://hlm.lge.com/qi'

#Release Version 1.6.4

startSP = 'TVSP16_1'
endSP = 'TVSP20_2'
updateSP = 'TVSP21_1'
filename = "Initiative일정관리_180611_V1"
logfilename = filename+"_log.txt"
openfilename = filename+".xlsx"
savefilename = filename+"_AutoUpdate.xlsx"
storyCalSupport = True #False
ID = "sungbin.na"
PASSWORD = ""

default_Sprint_Info = {
    'TVSP16_1' : '',  'TVSP16_2' : '',  'TVSP17_1' : '',  'TVSP17_2' : '',
    'TVSP18_1' : '',  'TVSP18_2' : '',  'TVSP19_1' : '',  'TVSP19_2' : '',
    'TVSP20_1' : '',  'TVSP20_2' : '',  'TVSP21_1' : '',  'TVSP21_2' : '',
    'TVSP22_1' : '',  'TVSP22_2' : '',  'TVSP23_1' : '',  'TVSP23_2' : '',
#    'TVSP24_1' : '',  'TVSP24_2' : '',  'TVSP25_1' : '',  'TVSP25_2' : '',
#    'TVSP26_1' : '',  'TVSP26_2' : '',  'TVSP27_1' : '',  'TVSP27_2' : '',
#    'TVSP28_1' : '',  'TVSP28_2' : '',  'TVSP29_1' : '',  'TVSP29_2' : '',
#    'TVSP30_1' : '',  'TVSP30_2' : '',  'TVSP31_1' : '',  'TVSP31_2' : '',
    }

default_epic_info = {
        'Epic Key' : '',
        'Release_SP' : '',
        'Summary' : "",
        'Assignee' : '',
        'duedate' : '',
        'Status' : '',
        'CreatedDate' : '',
        'TVSP' : default_Sprint_Info,
        'StoryCnt': 0,
        'StoryResolutionCnt' : 0,
        'RescheduleCnt' : 0,
        'STORY' : [],
        'AbnormalEpicSprint' : '',
        "GovOrDeployment" : '',
    }

default_initiative_info = {
    'Initiative Key' : '',
    'Summary' : '',
    'Assignee' : '',
    'Status' : '',
    'Release_SP' : '',
    'CreatedDate' : '',
    '관리대상' : '',
    'Risk 관리 대상' : '',
    'Initiative Order' : '',
    'EPIC' : [],
    'DEMO' : [],
    'CCC' : [],
    'TestCase' : [],
    'Dev_Verification' : [],
    'TVSP' : default_Sprint_Info,
    'Status Color' : '',
    'SE_Delivery' : '',
    'SE_Quality' : '',
    'ScopeOfChange' : '',
    'EpicCnt' : '',
    'EpicResolutionCnt' : '',
    'StoryCnt' : 0,
    'StoryResolutionCnt' : 0,
    'RMS' : '',
    'RescheduleCnt' : 0,
    'EpicDelayedCnt' : 0,
    'STEOnSite' : '',
    'AbnormalEpicSprint' : '',
    "GovOrDeployment" : '',
        }

ExceptionList = { }

#####################################################################################################################
# JIRA Control
maxResultCnt = 3000



#===========================================================================
# Get filtered issue with Filter ID in Dev Tracker
# filter 지정을 통한 Initiative 검색
#===========================================================================
def getFilterIDResult(jiraHandle, filterId, getFieldList=()) :
    setFilterID = 'filter = ' + str(filterId)
    print('strFilterID =', setFilterID)
    #resultIssue = jiraHandle.search_issues(setFilterID, startAt = 0, maxResults = 1000, fields = setfield, expand=None)
    resultIssue = jiraHandle.search_issues(setFilterID, startAt = 0, maxResults = maxResultCnt, fields = getFieldList, expand=None)
    print("[Tracker] Get JIRA Issue with Specific Filter ID: " + setFilterID)
    return resultIssue


#===========================================================================
# Get filtered issue with Filter ID in Dev Tracker
# Query 지정을 통한 Initiative 검색
#===========================================================================
def getFilterQueryResult(jiraHandle, filterQuery, getFieldList=None) :
    # Get Filtered issue with JQL Querfy String in Q/Dev Tracker
    #setFilter = 'Filter in (M3.LK61.EU.QA1, M3.LK61.EU.QA2, M3.LK61.EU.QA3, M3.LK61.EU.QA4)'
    resultIssue = jiraHandle.search_issues(filterQuery, startAt = 0, maxResults = maxResultCnt, fields = getFieldList, expand=None)
    #print("[Tracker] Get JIRA Issue with Specific Filter String: " + filterQuery)
    return resultIssue


#===========================================================================
# Get Key of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getKey(jiraIssue) :
    value = jiraIssue.raw['key']
    if(value != None) :
        #print("key = ", value)
        return value

    #print("key = Null")
    return None


#===========================================================================
# Get Summary of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getSummary(jiraIssue) :
    value = jiraIssue.raw['fields']['summary']
    if(value != None) :
        #print("summary = ", value)
        return value

    #print("summary = Null")
    #return None
    return 'None'


#===========================================================================
# Get Status of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getStatus(jiraIssue) :
    value = jiraIssue.raw['fields']['status']['name']
    #print("status = ", value)
    return value


#===========================================================================
# Get issuetype of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getIssuetype(jiraIssue) :
    value = jiraIssue.raw['fields']['issuetype']['name']
    #print("issuetype = ", value)
    return value


#===========================================================================
# Get resolution of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getResolution(jiraIssue) :
    value = jiraIssue.raw['fields']['resolution']
    #print("resolution = ", value)
    return value


#===========================================================================
# Get components of jira
# [param] jiraIssue : json object of jira
# [return] components[]
#===========================================================================
def getComponents(jiraIssue) :
    value = jiraIssue.raw['fields']['components']

    if(value != None) :
        #print("components = ", json.dumps(value))
        return value

    #print("components = Null")
    #return None
    return 'None'


#===========================================================================
# Get components of jira
# [param] jiraIssue : json object of jira
# [return] "O" or "X"
#===========================================================================
def checkGovDeployComponents(jiraIssue) :
    value = jiraIssue.raw['fields']['components']

    if(value != None) :
        for comp in value :
            #print("********* Comp = ", comp)
            if("Deployment(manage)" in comp['name']) :
                #print("[True] Governing or Deployment(manage) components = ", comp['name'])
                log.write("[True] Deployment(manage) components = {}".format(comp['name']))
                return "O"
            if("_Governing" in comp['name']) :
                #print("[True] Governing or Deployment(manage) components = ", comp['name'])
                log.write("[True] Governing components = {}".format(comp['name']))
                return "O"
    #print("components = Null")
    return "X"


#===========================================================================
# Get Release Sprint of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getReleaseSprint(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_15926']
    if(value != None) :
        #print("Release Sprint = ", value)
        retstr = "".join(value)
        return retstr

    #print("Release Sprint = Null")
    #return None
    return 'None'


#===========================================================================
# convert created date to datetime obect
# [param] createddate : created
# [return] datetime object
#===========================================================================
def conversionCreatedDateToDatetime(datedata) :
    # "2018-05-04"
    # "2018-05-04T14:03:49.000+0900"
    datedata = str(datedata)
    result = datedata.split(' ')
    result = result[0] # take yyyy-mm-dd area
    result = result.split('-')

    if(len(result) >= 3) :
        yyyy = result[0]
        mm = result[1]
        dd = result[2]
        dd = dd.split('T')
        dd = dd[0]
        return datetime(int(yyyy), int(mm), int(dd))
    else :
        return 'DueDate미설정'


#===========================================================================
# convert duedate to Sprint
# [param] duedate : duedate
# [return] Sprint str
#===========================================================================
def conversionDuedateToSprint(duedate) :
    targetDate = conversionCreatedDateToDatetime(duedate)

    #print("conversionDuedateToSprint = org {0}, conversion = {1}".format(duedate, targetDate))
    log.write("conversionDuedateToSprint = org {0}, conversion = {1}".format(duedate, targetDate))

    if(targetDate == 'DueDate미설정') :
        return 'TVSP_UNDEF'
    elif(targetDate < TVSP11_Start) :
        return 'TVSP11이전항목'
    if(targetDate >= TVSP11_Start and targetDate < TVSP12_Start) :
        return 'TVSP11'
    elif(targetDate >= TVSP12_Start and targetDate < TVSP13_Start) :
        return 'TVSP12'
    elif(targetDate >= TVSP13_Start and targetDate < TVSP14_Start) :
        return 'TVSP13'
    elif(targetDate >= TVSP14_Start and targetDate < TVSP15_Start) :
        return 'TVSP14'
    elif(targetDate >= TVSP15_Start and targetDate < TVSP16_Start) :
        return 'TVSP15'
    elif(targetDate >= TVSP16_Start and targetDate < TVSP17_Start) :
        return 'TVSP16'
    elif(targetDate >= TVSP17_Start and targetDate < TVSP18_Start) :
        return 'TVSP17'
    elif(targetDate >= TVSP18_Start and targetDate < TVSP19_Start) :
        return 'TVSP18'
    elif(targetDate >= TVSP19_Start and targetDate < TVSP20_Start) :
        return 'TVSP19'
    elif(targetDate >= TVSP20_Start and targetDate < TVSP21_Start) :
        return 'TVSP20'
    elif(targetDate >= TVSP21_Start and targetDate < TVSP22_Start) :
        return 'TVSP21'
    elif(targetDate >= TVSP22_Start and targetDate < TVSP23_Start) :
        return 'TVSP22'
    elif(targetDate >= TVSP23_Start and targetDate < TVSP24_Start) :
        return 'TVSP23'
    elif(targetDate >= TVSP24_Start and targetDate < TVSP25_Start) :
        return 'TVSP24'
    elif(targetDate >= TVSP25_Start and targetDate < TVSP26_Start) :
        return 'TVSP25'
    elif(targetDate >= TVSP26_Start and targetDate < TVSP27_Start) :
        return 'TVSP26'
    elif(targetDate >= TVSP27_Start and targetDate < TVSP28_Start) :
        return 'TVSP27'
    elif(targetDate >= TVSP28_Start and targetDate < TVSP29_Start) :
        return 'TVSP28'
    elif(targetDate >= TVSP29_Start and targetDate < TVSP30_Start) :
        return 'TVSP29'
    elif(targetDate >= TVSP30_Start and targetDate < TVSP31_Start) :
        return 'TVSP30'
    elif(targetDate >= TVSP31_Start and targetDate < TVSP32_Start) :
        return 'TVSP31'
    elif(targetDate >= TVSP32_Start) :
        return 'TVSP32이후항목'

#===========================================================================
# convert Sprint to duedate
# [param] duedate : Sprint
# [return] Datetime
#===========================================================================
def conversionSprnitToDatetime(Sprint) :

    if(Sprint == 'TVSP11이전항목') :
        result = datetime(2018, 1, 14)
    elif(Sprint == 'TVSP11') :
        result = datetime(2018, 1, 28)
    elif(Sprint == 'TVSP12') :
        result = datetime(2018, 2, 11)
    elif(Sprint == 'TVSP13') :
        result = datetime(2018, 2, 25)
    elif(Sprint == 'TVSP14') :
        result = datetime(2018, 3, 11)
    elif(Sprint == 'TVSP15') :
        result = datetime(2018, 4, 1)
    elif(Sprint == 'TVSP16') :
        result = datetime(2018, 4, 15)
    elif(Sprint == 'TVSP17') :
        result = datetime(2018, 4, 29)
    elif(Sprint == 'TVSP18') :
        result = datetime(2018, 5, 13)
    elif(Sprint == 'TVSP19') :
        result = datetime(2018, 5, 27)
    elif(Sprint == 'TVSP20') :
        result = datetime(2018, 6, 10)
    elif(Sprint == 'TVSP21') :
        result = datetime(2018, 6, 24)
    elif(Sprint == 'TVSP22') :
        result = datetime(2018, 7, 8)
    elif(Sprint == 'TVSP23') :
        result = datetime(2018, 7, 22)
    elif(Sprint == 'TVSP24') :
        result = datetime(2018, 8, 5)
    elif(Sprint == 'TVSP25') :
        result = datetime(2018, 8, 19)
    elif(Sprint == 'TVSP26') :
        result = datetime(2018, 9, 2)
    elif(Sprint == 'TVSP27') :
        result = datetime(2018, 9, 16)
    elif(Sprint == 'TVSP28') :
        result = datetime(2018, 9, 30)
    elif(Sprint == 'TVSP29') :
        result = datetime(2018, 10, 14)
    elif(Sprint == 'TVSP30') :
        result = datetime(2018, 10, 28)
    elif(Sprint == 'TVSP31') :
        result = datetime(2018, 11, 11)
    elif(Sprint == 'TVSP32') :
        result = datetime(2018, 11, 25)
    elif(Sprint == 'TVSP32이후') :
        result = datetime(2018, 12, 31)
    else : #TVSP_UNDEF
        result = datetime(2018, 12, 31)

    log.write("Sprint = {0}, conversionSprnitToDatetime = {1}, ".format(Sprint, result))
    return result

#===========================================================================
# convert ReleaseSprint to ShortSprint
# [param] duedate : duedate
# [return] Sprint str
#===========================================================================
def conversionReleaseSprintToSprint(ReleaseSprint) :
    bypass = False
    #for sprint in ReleaseSprint :
    sprint = ReleaseSprint
    if( "GL2_" in sprint) :
        #print(sprint)
        a = sprint.replace('GL2_', '')
        a = a.split('_')
        b = a[0]
    elif ("FC2_" in sprint) :
        b = sprint
    elif ("19Y_" in sprint) :
        b = sprint
    else :
        bypass = True

    if(bypass == True) :
        #print(sprint)
        return sprint
    else :
        b = b.replace('IR1', '')
        b = b.replace('IR2', '')
        b = b.replace('IR3', '')
        b = b.replace('IR4', '')
        #print(b)
        b = b.split('(')
        #print(b[0])
        return b[0]


#===========================================================================
# Get Status Summary of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getStatusSummary(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_15710']
    if(value != None) :
        #print("Status Summary = ", value)
        return value

    #print("Status Summary = Null")
    #return None
    return 'None'


#===========================================================================
# Get Status Color of jira
# [param] jiraIssue : json object of jira
# [return] str (RGB)
#===========================================================================
def getStatusColor(jiraIssue) :
    statusColor = jiraIssue.raw['fields']['customfield_15711']
    if(statusColor != None) :
        #print("Status Color = ", json.dumps(statusColor['value']))
        return statusColor['value']

    #print("Status Color = Null")
    #return None
    return 'None'


#===========================================================================
# Get SE_Delivery of jira
# [param] jiraIssue : json object of jira
# [return] str (RGB)
#===========================================================================
def getSE_Delivery(jiraIssue) :
    delivery = jiraIssue.raw['fields']['customfield_16988']
    if(delivery != None) :
        #print("SE_Delivery = ", json.dumps(delivery))
        return delivery['value']

    #print("SE_Delivery = Null")
    #return None
    return 'None'


#===========================================================================
# Get SE_Quality of jira
# [param] jiraIssue : json object of jira
# [return] str (RGB)
#===========================================================================
def getSE_Quality(jiraIssue) :
    quality = jiraIssue.raw['fields']['customfield_16987']
    if(quality != None) :
        #print("SE_Quality = ", json.dumps(quality))
        return quality['value']

    #print("SE_Quality = Null")
    #return None
    return 'None'

#===========================================================================
# Get D_Comment of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getD_Comment(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_16984']
    #print("D_Comment = ", value)
    return value


#===========================================================================
# Get Q_Comment of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getQ_Comment(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_16983']
    #print("Q_Comment = ", value)
    return value


#===========================================================================
# Get STE Member List of jira
# [param] jiraIssue : json object of jira
# [return] QE[]
#===========================================================================
def getSTEList(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_15228']
    if(value != None) :
        #print("STE_List[] = ", json.dumps(value))
        return value

    #print("STE_List[] = Null")
    #return None
    return 'None'


#===========================================================================
# Get Initiative Order of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getInitiativeOrder(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_16986']
    #print("Initiative Order = ", value)
    return value


#===========================================================================
# Get Initiative Score of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getInitiativeScore(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_16985']
    #print("Initiative Score = ", value)
    return value

#===========================================================================
# Get Created Date of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getCreatedDate(jiraIssue) :
    value = jiraIssue.raw['fields']['created']
    #print("Created Date = ", value)
    return value


#===========================================================================
# Get Updated Date of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getUpdatedDate(jiraIssue) :
    value = jiraIssue.raw['fields']['updated']
    #print("Updated Date = ", value)
    return value


#===========================================================================
# Get Due Date of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getDueDate(jiraIssue) :
    value = jiraIssue.raw['fields']['duedate']
    #print("Due Date = ", value)
    return value

#===========================================================================
# Get Resolution Date List of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getResolutionDate(jiraIssue) :
    value = jiraIssue.raw['fields']['resolutiondate']
    #print("Resolutiondate Date = ", value)
    return value


#===========================================================================
# Get Created Date List of jira
# [param] jiraIssue : json object of jira
# [return] labels []
#===========================================================================
def getLabels(jiraIssue) :
    value = jiraIssue.raw['fields']['labels']
    #print("labels = ", value)
    return value


#===========================================================================
# Get Description List of jira
# [param] jiraIssue : json object of jira
# [return] labels [ 'a', 'b', .... ]
#===========================================================================
def getDescription(jiraIssue) :
    value = jiraIssue.raw['fields']['description']
    #print("description = ", value)
    return value

#===========================================================================
# Get fixVersions of jira
# [param] jiraIssue : json object of jira
# [return] fixVersions [ { 'name' : '' } ]
#===========================================================================
def getFixVersions(jiraIssue) :
    value = jiraIssue.raw['fields']['fixVersions']
    #print("fixVersions = ", value)
    return value


#===========================================================================
# Get Scope of Change of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getScopeOfChange(jiraIssue) :
    value = jiraIssue.raw['fields']['customfield_15104']
    if(value != None) :
        #print("Scope of Change = ", value['value'])
        return value['value']

    #print("Scope of Change = Null")
    #return None
    return 'None'


#===========================================================================
# Get Issue Links of jira
# [param] jiraIssue : json object of jira
# [return] issuelinks[ {}, .... ]
#===========================================================================
def getIssueLinks(jiraIssue) :
    value = jiraIssue.raw['fields']['issuelinks']
    #print("Issue Links = ", value)
    return value

#===========================================================================
# Get Reporter of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getReporter(jiraIssue) :
    value = jiraIssue.raw['fields']['reporter']['name']
    #print("reporter = ", value)
    return value


#===========================================================================
# Get Assignee of jira
# [param] jiraIssue : json object of jira
# [return] str
#===========================================================================
def getAssignee(jiraIssue) :
    value = "Unassigned"
    if(jiraIssue.raw['fields']['assignee'] != None) :
        value = jiraIssue.raw['fields']['assignee']['name']
        return value
    #print("assignee = ", value)
    return value


#===========================================================================
# Get Reporter of jira
# [param] jiraIssue : json object of jira
# [return] Watchers [ ] <== [ { 'name' : ''}, { 'emailAddress' : '' }, .... ]
#===========================================================================
def getWatchers(jiraIssue) :
    watchers = jiraIssue.raw['fields']['customfield_10105']
    results = []
    for watcher in watchers :
        results.append(watcher)

    #print("Watcher List = ", results)
    return results

#===========================================================================
# Get Epics / Milestone block from Description of jira
# [param] description : description string
# [param] fieldtitle : Title of block like '*Milestone*'
# [return] str or None
#===========================================================================
strEpic = ["*개발 Epic 산정*", "*Epics*", "*EPICs*" ]
strMilestone = [ "*Milestone", "*Expected Deliveries*" ]

def getEpicsMilestoneFromDesc(description, fieldtitle) :
    startpos = description.find(fieldtitle)
    if(startpos > 1) :
        endpos = description.find('*', startpos + len(fieldtitle))

    if(startpos >= 1 and endpos > startpos):
        result = description[startpos:endpos]
        #print(result)
        return result

    #print("getEpicsFromDesc = Null")
    #return None
    return 'None'

#===========================================================================
# Check STE On Site Initiative or Not
# [param] dissue : jira issue
# [return] True (STE On Site) or False (None)
#===========================================================================
def checkLabels(dissue, labelname) :
    labels = getLabels(dissue)
    for label in labels :
        if(label == labelname) :
            return "O"
    return "X"

#####################################################################################################################
# Exel Control

#===========================================================================
# Get row count of excel (with Data)
# [param] Sheetname : Excel Sheet handle
# [param] rowpos : start row position of Data ( 1 ~ XXXX )
# [param] colpos : reference col index to detect exact row count (This column should be filled with data)
# [return] row count
#===========================================================================
def getRowCount(Sheetname, rowpos, colpos) :
    for i in range(rowpos, 50000) :
        val = Sheetname.cell(row = i, column = colpos).value
        if(val == None) :
            break;

    print("Row Count of ", Sheetname, " = ", i-rowpos)
    return (i-rowpos)


#===========================================================================
# Get Column count of excel (with Data)
# [param] Sheetname : Excel Sheet handle
# [param] rowpos : start row position of Title / Header
# [param] colpos : start col position of title
# [return] column count
#===========================================================================
def getColumnCount(Sheetname, rowpos, colpos) :
    for i in range(colpos,50000) :
        val = Sheetname.cell(row = rowpos, column = i).value
        if(val == None) :
            break;

    print("Column Count of ", Sheetname, " = ", i-colpos)
    return (i-colpos)



#===========================================================================
# Get Column Index of excel (with title)
# [param] Sheetname : Excel Sheet handle
# [param] rowpos : start row position of Title / Header
# [return] column Index
#===========================================================================
def getColumnIndex(Sheetname, row, title) :
    index = 1
    for col in range(1, Sheetname.max_column) :
        if(title == str(Sheetname.cell(row = row, column = col).value)) :
            break;
        index += 1

    if (Sheetname.max_column <= index) :
        index = None
        #print("title = ", title, " : Can't find title in exel")
    else :
        #print("title = ", title, ", Index = ", index)
        pass

    return (index)



#===========================================================================
# Get row Index of initiative to find
# [param] Sheetname : Excel Sheet handle
# [param] InitiativeKey : InitiativeKey to find
# [return] row Index or 0 (Not found)
#===========================================================================
def getInitiativeRowIndex(Sheetname, InitiativeKey) :
    isFound = False
    result = 0
    type = 0
    keyvalue = 0
    #print("getInitiativeRowIndex = Max_RowCount = ", MAX_RowCount)
    for row_index in range(3, MAX_RowCount+1) :
        type = str(Sheetname.cell(row = row_index, column = 5).value).strip()
        keyvalue = str(Sheetname.cell(row = row_index, column = 6).value).strip()
        #print("######>>>>> row_index = {3}, Issue Type = {0}, KeyValue = {1}, Initiative Key = {2}".format(type, keyvalue, InitiativeKey, row_index))
        if(type == "Initiative") :
            keyvalue = str(Sheetname.cell(row = row_index, column = 6).value).strip()
            if(keyvalue == InitiativeKey.strip()) :
                isFound = True
                result = row_index
                #print("Found Initiative Key of ", Sheetname, " : Key = ", InitiativeKey, " keyvalue = ", keyvalue, " Index = ", row_index)
                break;

    if(isFound == False) :
        #print("Not Found Initiative Key of ", Sheetname, " : Key = ",InitiativeKey, " keyvalue = ", keyvalue, " Index = ", result, " Type = ", type)
        result = 0

    return result




#===========================================================================
# Get row Index of Epic to find
# [param] Sheetname : Excel Sheet handle
# [param] InitiativeKey : Epic to find
# [return] row Index or 0 (Not found)
#===========================================================================
def getEpicRowIndex(Sheetname, EpicKey) :
    isFound = False
    result = 0
    type = 0
    #print("getEpicRowIndex = Max_RowCount = ", MAX_RowCount)
    for row_index in range(3, MAX_RowCount+1) :
        type = str(Sheetname.cell(row = row_index, column = CI_IssueType).value).strip()
        if(type == "EPIC") :
            keyvalue = str(Sheetname.cell(row = row_index, column = CI_EpicKey).value).strip()
            if(keyvalue == EpicKey) :
                isFound = True
                result = row_index
                print("Found Epic Key of ", Sheetname, " : Key = ", EpicKey, ", Index = ", row_index)
                break;

    if(isFound == False) :
        print("Not Found Epic Key of ", Sheetname, " : Key = ", EpicKey, ", Index = ", result)
        result = 0

    return result


#===========================================================================
# Get title list of header / title
# [param] Sheetname : Excel Sheet handle
# [param] row : start row position of Title / Header
# [param] col : start col position of title
# [return] title[]
#===========================================================================
def getTitleListfromXls(Sheetname, row, col) :
    title = []
    for i in range(col, MAX_ColCount+1) :
        title.append(str(Sheetname.cell(row = row, column = i).value).strip())

    #print("Title List of ", Sheetname, " = ", title)
    return title



#===========================================================================
# Get All Initiative Key List from Excel
# [param] Sheetname : Excel Sheet handle
# [param] rowpos : start row position of Data ( 1 ~ XXXX )
# [return] Initiative Key List[ 'KEY1', 'KEY2', ... ]
#===========================================================================
def getInitiativeKeylistFromXls(Sheetname, row) :
    initative_key = []

    #print("getInitiativeKeylistFromXls = Max_RowCount = ", MAX_RowCount)
    for row_index in range(row, MAX_RowCount+1) :
        type = str(Sheetname.cell(row = row_index, column = CI_IssueType).value).strip()
        if(type == "Initiative") :
            initative_key.append(str(Sheetname.cell(row = row_index, column = CI_InitKey).value).strip())

    #print("Initiative Key List of ", Sheetname, " = ", initative_key)
    return initative_key


#===========================================================================
# Get All Epic Key List from Excel
# [param] Sheetname : Excel Sheet handle
# [param] rowpos : start row position of Data ( 1 ~ XXXX )
# [return] Epic Key List[ 'KEY1', 'KEY2', ... ]
#===========================================================================
def getEpicKeyListfromXls(Sheetname, row) :
    epic_key = []
    #print("getEpicKeyListfromXls = Max_RowCount = ", MAX_RowCount)
    for row_index in range(row, MAX_RowCount+1) :
        type = str(Sheetname.cell(row = row_index, column = CI_IssueType).value).strip()
        if(type == "EPIC") :
            epic_key.append(str(Sheetname.cell(row = row_index, column = CI_EpicKey).value).strip())

    epic_key = RemoveDuplicateInList(epic_key)
    #print("Epic Key List of ", Sheetname, " = ", epic_key)
    return epic_key



#===========================================================================
# Get Sprint History (dict) from Excel
# [param] Sheetname : Excel Sheet handle
# [param] KeyID : Initiative or Epic Key to get Sprint History from excel
# [return] Sprint_Info{ 'SP16_1' : '', 'SP16_2' : '', .... }
#===========================================================================
def getSprintHistoryfromXls(Sheetname, KeyID, IssueType, rowIndex) :
    Sprint_Info = { }
    Sprint_Info = copy.deepcopy(default_Sprint_Info)

    if(rowIndex == 0) :
        pass
    else :
        for col_index in range(CI_StartPos, CI_EndPos+1) :
            tvspstr = str(Sheetname.cell(row = 2, column = col_index).value).strip()
            Sprint_Info[tvspstr] = str(Sheetname.cell(row = rowIndex, column = col_index).value).strip()

        if(IssueType == "Initiative") :
            #print("===============Update Initiative Sprint_Info===================== Row Index =", rowIndex)
            pass
        elif (IssueType == "EPIC") :
            #print("===============Update Epic Sprint_Info===================== Row Index =", rowIndex)
            pass
    #print (Sprint_Info)
    return Sprint_Info



#===========================================================================
# Get Epic Information from Excel
# [param] Sheetname : Excel Sheet handle
# [param] EpicKey : Epic Key to get detail information from excel
# [return] epic_info { 'key' : '', 'summary' : '', Sprint_Info{ 'SP16_1' : '', 'SP16_2' : '', .... } } or None
#===========================================================================
def getEpicInfofromXls(Sheetname, EpicKey) :
    epic_info = { }
    bfound = False
    getEpicKey = 0
    epic_info = copy.deepcopy(default_epic_info)
    #print("getEpicInfofromXls = Max_RowCount = ", MAX_RowCount)
    for row_index in range(1, MAX_RowCount+1) :
        getEpicKey = str(Sheetname.cell(row = row_index, column = CI_EpicKey).value).strip()

        if(getEpicKey == EpicKey) :
            #print("&&&&&&&&&&&&&&&  getEpicKey = {0} == EpicKey = {1}".format(getEpicKey, EpicKey))
            #log.write("\n&&&&&&&&&&&&&&&  getEpicKey = {0} == EpicKey = {1}".format(getEpicKey, EpicKey))
            epic_info['Epic Key'] = getEpicKey
            epic_info['Release_SP'] = str(Sheetname.cell(row = row_index, column = CI_ReleaseSP).value).strip()
            spInfo = getSprintHistoryfromXls(Sheetname, getEpicKey, "EPIC", row_index)
            epic_info['TVSP'] = spInfo
            bfound = True

    if(bfound == True) :
        #print(epic_info)
        #log.write("\n@@ Epic Key = {0}\n".format(EpicKey))
        pass
    else :
        epic_info['Epic Key'] = EpicKey

        #log.write("\n@@ Not found Epic Key from xls = {0}\n".format(EpicKey))

    log.write(str(epic_info))
    return epic_info



#===========================================================================
# Get All Initiative - Epic Lists from Excel
# [param] Sheetname : Excel Sheet handle
# [return] epic_key [ { 'key' : 'Initative Key',  'epiclist' : [ 'Epic Key1', 'Epic Key2', ... ]}, ....  }
#===========================================================================
def getInitiativeAllEpicsListfromXls(Sheetname) :
    epic_key = []
    keylist = getInitiativeKeylistFromXls(Sheetname, 3)
    #print("getInitiativeAllEpicsListfromXls = Max_RowCount = ", MAX_RowCount)

    for keyID in keylist :
        tmp = { 'key' : '', 'epiclist' : []}
        tmp['key'] = keyID
        for row_index in range(1, MAX_RowCount+1) :
            type = str(Sheetname.cell(row = row_index, column = CI_IssueType).value).strip()
            epicparent = str(Sheetname.cell(row = row_index, column = CI_InitKey).value).strip()
            if(type == 'EPIC' and epicparent == keyID) :
                tmp['epiclist'].append(str(Sheetname.cell(row = row_index, column = CI_EpicKey).value).strip())
        epic_key.append(tmp)

    #print("*********** All Epic key List from Xls **********************")
    #print(epic_key)
    return epic_key


#===========================================================================
# Get a specific Initiative - Epic Lists from Excel
# [param] Sheetname : Excel Sheet handle
# [param] InitiativeKey : Initiative Key to get Epic Lists
# [return] EpicList[ 'Epic Key1', 'Epic Key2', ... ]
#===========================================================================
def getInitiativeEpicListsfromXls(Sheetname, InitiativeKey) :
    epic_key = []
    #print("getInitiativeEpicListsfromXls = Max_RowCount = ", MAX_RowCount)
    for row_index in range(1, MAX_RowCount+1) :
        type = str(Sheetname.cell(row = row_index, column = CI_IssueType).value).strip()
        getInitkey = str(Sheetname.cell(row = row_index, column = CI_InitKey).value).strip()
        if(type == "EPIC" and InitiativeKey == getInitkey) :
            epic_key.append(str(Sheetname.cell(row = row_index, column = CI_EpicKey).value).strip())

    #print("Initiative Key = ", InitiativeKey, " Epic Key List of ", Sheetname, " = ", epic_key)
    return epic_key



#===========================================================================
# Set value to Cell
# [param] Sheetname : Excel Sheet handle
# [param] row : row
# [param] col : column
# [return] None
#===========================================================================
def setXlsCell(Sheetname, row, col, value, format, link) :
    if(format == True) :
        Sheetname.cell(row=row, column=col).font = myfont
        Sheetname.cell(row=row, column=col).alignment = myalignment
        Sheetname.cell(row=row, column=col).fill = myfill
        Sheetname.cell(row=row, column=col).border = myborder

    if(link != None) :
        Sheetname.cell(row=row, column=col).hyperlink = link

    if(value != None) :
        Sheetname.cell(row=row, column=col).value = value


#===========================================================================
# Get value from Cell
# [param] Sheetname : Excel Sheet handle
# [param] row : row
# [param] col : column
# [return] value
#===========================================================================
def getXlsCell(Sheetname, row, col) :
    value = str(Sheetname.cell(row, col).value).strip()
    return value



#===========================================================================
# Get All Initiative Key List from Jira
# [param] rowpos : start row position of Data ( 1 ~ XXXX )
# [return] Initiative Key List[ 'KEY1', 'KEY2', ... ]
#===========================================================================
def getInitiativeKeylistFromJira(filterResult) :
    initiative_key = []
    for issue in filterResult :
        initiative_key.append(getKey(issue))

    #print("Initiative Key List from Jira = ", initative_key)
    return initiative_key



#===========================================================================
# Get All Initiative - Epic Lists from Jira
# [param] filterResult : Jira Result from Filtered JIRA Query
# [return] list[ { 'key' : 'Initative Key',  'epiclist' : [ 'Epic Key1', 'Epic Key2', ... ]}, ....  ]
#===========================================================================
def getInitiativeAllEpicsListfromJira(filterResult) :
    epic_key = []

    for dissue in filterResult :
        # Get issue with All Fields in Dev Tracker
        tmp = { 'key' : '', 'epiclist' : []}
        tmp['key'] = dissue.raw['key']
        bfound = False
        for issuelink in dissue.raw['fields']['issuelinks'] :
            if 'outwardIssue' in issuelink :
                if(issuelink['outwardIssue']['fields']['issuetype']['name'] == 'Epic') :
                    #print ("Key = ", dissue.raw['key'], " Status = ", issuelink['outwardIssue']['fields']['status']['name'], " Linked Issue = ", issuelink['outwardIssue']['key'])
                    tmp['epiclist'].append(issuelink['outwardIssue']['key'])
                    bfound = True
            if 'inwardIssue' in issuelink :
                if(issuelink['inwardIssue']['fields']['issuetype']['name'] == 'Epic') :
                    #print ("Key = ", dissue.raw['key'], " Status = ", issuelink['inwardIssue']['fields']['status']['name'], " Linked Issue = ", issuelink['inwardIssue']['key'])
                    tmp['epiclist'].append(issuelink['inwardIssue']['key'])
                    bfound = True

        if(bfound == True) :
            epic_key.append(tmp)

    #print("*********** Initiative - All Epic key List **********************")
    #print(epic_key)
    return epic_key



#===========================================================================
# Get All Epic Lists from Jira
# [param] filterResult : Jira Result from Filtered JIRA Query (Initiative or Epic Filter)
# [return] epic_key[ 'Epic Key1', 'Epic Key2', ... ]
#===========================================================================
def getEpicKeyListfromJira(filterResult, rawData) :
    epic_key = []

    if(rawData == "Initiative_Filter") : # make a epic list from Issuelinks
        for dissue in filterResult :
            # Get issue with All Fields in Dev Tracker
            for issuelink in dissue.raw['fields']['issuelinks'] :
                if 'outwardIssue' in issuelink :
                    if(issuelink['outwardIssue']['fields']['issuetype']['name'] == 'Epic') :
                        #print ("Key = ", dissue.raw['key'], " Status = ", issuelink['outwardIssue']['fields']['status']['name'], " Linked Issue = ", issuelink['outwardIssue']['key'])
                        epic_key.append(issuelink['outwardIssue']['key'])
                if 'inwardIssue' in issuelink :
                    if(issuelink['inwardIssue']['fields']['issuetype']['name'] == 'Epic') :
                        #print ("Key = ", dissue.raw['key'], " Status = ", issuelink['inwardIssue']['fields']['status']['name'], " Linked Issue = ", issuelink['inwardIssue']['key'])
                        epic_key.append(issuelink['inwardIssue']['key'])
    elif (rawData == "Epic_Filter") : # make a epic list from Epic Filter Result
        # Compare Epic List ........
        for dissue in filterResult :
            epic_key.append(getKey(dissue))
    else :
        pass

    # remove duplicate item in list
    epic_key = RemoveDuplicateInList(epic_key)

    #print("*********** All Epic key List from Jira (rawData = {0})**********************".format(rawData))
    #print(epic_key)
    return epic_key



#===========================================================================
# Get All Epic - Story and Task Lists from Jira
# [param] filterResult : Jira Result from Filtered JIRA Query (Epic Filter)
# [return] Story key[ 'Story Key1', 'Story Key2', 'Task Key3'... ]
#===========================================================================
def getStorynTaskKeyListfromJira(storyissue) :
    story_key = []
    for story in storyissue :
        story_key.append(story.raw['key'])

    #print("*********** All Story key List from Jira (dissue.raw['key'] = {0})**********************".format(storyissue.raw['key']))
    #print(story_key)
    return story_key

#===========================================================================
# Get Epic Resolved Count of Initiative from Jira
# [param] filterResult : Jira Result from Filtered JIRA Query (Initiative or Epic Filter)
# [return] epic_key[ 'Epic Key1', 'Epic Key2', ... ]
#===========================================================================
def getChildEpicResolvedCntfromJira(dissue) :
    resolvedCnt = 0
    # Get issue with All Fields in Dev Tracker
    for issuelink in dissue.raw['fields']['issuelinks'] :
        if 'outwardIssue' in issuelink :
            if(issuelink['outwardIssue']['fields']['issuetype']['name'] == 'Epic') :
                status = issuelink['outwardIssue']['fields']['status']['name']
                if(status == "Resolved" or status == "Closed" or status == "Deferred" or status == "Delivered") :
                    resolvedCnt += 1
        if 'inwardIssue' in issuelink :
            if(issuelink['inwardIssue']['fields']['issuetype']['name'] == 'Epic') :
                status = issuelink['inwardIssue']['fields']['status']['name']
                if(status == "Resolved" or status == "Closed" or status == "Deferred" or status == "Delivered") :
                    resolvedCnt += 1

    return resolvedCnt


#===========================================================================
# Get Story and Task Resolved Count of Initiative from Jira
# [param] filterResult : Jira Result from Filtered JIRA Query (Initiative or Epic Filter)
# [return] resolved count of Story or Task
#===========================================================================
def getChildStorynTaskResolvedCntfromJira(storyissue) :
    resolvedCnt = 0
    if(storyissue != None) :
        resolvedCnt = 0
        for story in storyissue :
            status = getStatus(story)
            if(status == "Resolved" or status == "Closed" or status == "Deferred" or status == "Delivered" or status == "Verify") :
                resolvedCnt += 1

    return resolvedCnt


#===========================================================================
# Get a specific Initiative - Epic Lists from Jira
# [param] jiraAllEpicList : list[ { 'key' : 'Initative Key',  'epiclist' : [ 'Epic Key1', 'Epic Key2', ... ]}, ....  }
# [param] InitiativeKey : Initiative Key to get Epic Lists
# [return] EpicList[ 'Epic Key1', 'Epic Key2', ... ]
#===========================================================================
def getInitiativeEpicListsfromJira(jiraAllEpicList, InitiativeKey) :

    for item in jiraAllEpicList :
        if (InitiativeKey == item['key']) :
            return item['epiclist']

    #print("*********** Initiative Key = {0} Epic key from JIRA **********************".format(InitiativeKey))
    #print(epic_key)
    return []



#===========================================================================
# Get the detail Initiative Information needed for history management from excel
# [param] Sheetname : Excel Sheet handle
# [param] IntiativeKeyList : Initiative List[ 'KEY1', 'KEY2', ... ]
# [param] Init_EpicList : Initative - Epic Info list[ { 'key' : 'Initative Key',  'epiclist' : [ 'Epic Key1', 'Epic Key2', ... ]}, ....  ]
# [return] all initiative_info list []
#===========================================================================
def getInitiativeDetailInfofromXls(Sheetname, IntiativeKeyList, Init_EpicList) :
    result = []
    initiative_info = {}
    for key in IntiativeKeyList :
        initiative_info = copy.deepcopy(default_initiative_info)
        rowIndex = getInitiativeRowIndex(Sheetname, key)

        if(rowIndex > 0) :
            #print("\n######## {0} row - Update Initiative Detail information from Xls".format(rowIndex))
            initiative_info["Initiative Key"] = str(Sheetname.cell(row = rowIndex, column = CI_InitKey).value).strip()
            initiative_info["Release_SP"] = str(Sheetname.cell(row = rowIndex, column = CI_ReleaseSP).value).strip()
            #initiative_info["관리대상"] = str(Sheetname.cell(row = rowIndex, column = CI_SPE_M).value).strip()
            #initiative_info["Risk 관리 대상"] = str(Sheetname.cell(row = rowIndex, column = CI_SPE_R).value).strip()
            #initiative_info["RMS"] = str(Sheetname.cell(row = rowIndex, column = CI_RMS).value).strip()

            #SP
            spInfo = getSprintHistoryfromXls(Sheetname, key, "Initiative", rowIndex)
            initiative_info['TVSP'] = spInfo

            #EPIC
            epic_list = getInitiativeEpicListsfromJira(Init_EpicList, key)
            #print("<<<<<<<<<< Initiative Key = {0}, Epic List = {1} >>>>>>>>>>>>>>>>>".format(key, epic_list))
            for epickey in epic_list :
                # epic_info = { 'Epic Key' : '', 'Summary' : '', Sprint_Info{ 'SP16_1' : '', 'SP16_2' : '', .... } }
                epicInfo = getEpicInfofromXls(Sheetname, epickey)
                if(epicInfo != None) :
                    initiative_info['EPIC'].append(epicInfo)

            result.append(initiative_info)
            #print(initiative_info)
        #20180516 : else: 문은 bug fix - Excel에 Initative key가 없으나 JIRA상에 신규로 생성된 Initiative가 존재 하는 경우 신규 추가를 위해 필요함.
        else :
            initiative_info["Initiative Key"] = key
            #EPIC
            epic_list = getInitiativeEpicListsfromJira(Init_EpicList, key)
            #print("<<<<<<<<<< Initiative Key = {0}, Epic List = {1} >>>>>>>>>>>>>>>>>".format(key, epic_list))
            for epickey in epic_list :
                # epicInfo = { 'Epic Key' : '', 'Summary' : '', Sprint_Info{ 'SP16_1' : '', 'SP16_2' : '', .... } }
                epicInfo = copy.deepcopy(default_epic_info)
                epicInfo['Epic Key'] = epickey
                initiative_info['EPIC'].append(epicInfo)

            result.append(initiative_info)

    return result



#===========================================================================
# Get the detail Initiative Information needed for history management from Jira
# [param] finalinfo : Initative Detail Info after updating data from excel.
#   final_info = [ default_initiative_info1, default_initiative_info2, default_initiative_info3, ..., default_initiative_infoN ]
#   default_initiative_info = {
#        'Initiative Key' : '',
#        'Summary' : '',
#        'Assignee' : '',
#        'Status' : '',
#        'Release_SP' : '',
#        'CreatedDate' : '',
#        '관리대상' : '',
#        'Risk 관리 대상' : '',
#        'Initiative Order' : '',
#        'EPIC' : [],
#        'DEMO' : [],
#        'CCC' : [],
#        'TestCase' : [],
#        'Dev_Verification' : [],
#        'TVSP' : {},
#        }
#
#    default_epic_info = {
#            'Epic Key' : '',
#            'Release_SP' : '',
#            'Summary' : "",
#            'Assignee' : '',
#            'duedate' : '',
#            'Status' : '',
#            'CreatedDate' : '',
#            'TVSP' : { },
#        }
#
# [return] Final Data : all initiative_info list []
#===========================================================================
def getInitiativeDetailInfofromJira(Initiative_FilterResult, Epic_FilterResult, finalinfo) :
    resultDB = []

    for initiative in finalinfo :
        dissue = getJiraInfo(Initiative_FilterResult, initiative['Initiative Key'])
        #print("#######>>> Initative Key = ", initiative['Initiative Key'])
        log.write("\n#######>>> Initative Key = {0} , EpicCnt = {1}".format(initiative['Initiative Key'], len(initiative['EPIC'])))
        if(dissue) :
            initiative['Summary'] = getSummary(dissue)
            initiative['Assignee'] = getAssignee(dissue)
            initiative['Status'] = getStatus(dissue)
            initiative['CreatedDate'] = getCreatedDate(dissue)
            initiative['Initiative Order'] = getInitiativeOrder(dissue)
            initiative['Status Color'] = getStatusColor(dissue)
            initiative['SE_Delivery'] = getSE_Delivery(dissue)
            initiative['SE_Quality'] = getSE_Quality(dissue)
            initiative['ScopeOfChange'] = getScopeOfChange(dissue)
            if(initiative['Status'] == "Delivered" or initiative['Status'] == "Deferred") :
                initiative['TVSP'][updateSP] = conversionDuedateToSprint(getResolutionDate(dissue)) #duedate 기반 SP 정보 기입 (Release Sprint 값 적용)
            else :
                initiative['TVSP'][updateSP] = conversionReleaseSprintToSprint(getReleaseSprint(dissue)) #duedate 기반 SP 정보 기입 (Release Sprint 값 적용)
            initiative['EpicCnt'] = len(initiative['EPIC'])
            initiative['EpicResolutionCnt'] = getChildEpicResolvedCntfromJira(dissue)
            initiative['StoryCnt'] = 0
            initiative['StoryResolutionCnt'] = 0
            initiative['EpicDelayedCnt'] = 0
            initiative['STEOnSite'] = checkLabels(dissue, "STE확인필요")
            initiative['관리대상'] = checkLabels(dissue, "SPE_M")
            initiative['Risk 관리 대상'] = checkLabels(dissue, "SPE_R")
            initiative['ScopeOfChange'] = getScopeOfChange(dissue)
            initiative['GovOrDeployment'] = checkGovDeployComponents(dissue)
            initative_due = conversionSprnitToDatetime(initiative['TVSP'][updateSP])

            for epic in initiative['EPIC'] :
                #print("#######>>> Epic Key = ", epic['Epic Key'])
                log.write("\n#######>>> Epic Key = {0} ".format(epic['Epic Key']))
                epicissue = getJiraInfo(Epic_FilterResult, epic['Epic Key'])
                if(epicissue) :
                    epic['Summary'] = getSummary(epicissue)
                    epic['Assignee'] = getAssignee(epicissue)
                    epic['Status'] = getStatus(epicissue)
                    epic['CreatedDate'] = getCreatedDate(epicissue)
                    epic['duedate'] = getDueDate(epicissue)
                    epic['GovOrDeployment'] = checkGovDeployComponents(epicissue)

                    #Epic DueDate미설정 항목은 미설정, Deliverd된 항목은 Resolution Date 기준으로 금주 Sprint 일정 Write
                    if(epic['Status'] == "Delivered" or epic['Status'] == "Resolved" or epic['Status'] == "Closed" or epic['Status'] == "Deferred") :
                        epic['TVSP'][updateSP] = conversionDuedateToSprint(getResolutionDate(epicissue))
                    else :
                        epic['TVSP'][updateSP] = conversionDuedateToSprint(getDueDate(epicissue))
                        due = conversionCreatedDateToDatetime(epic['duedate'])
                        if(due == 'DueDate미설정') :
                            initiative['EpicDelayedCnt'] += 1
                        else :
                            if(due < datetime.now()) :
                                initiative['EpicDelayedCnt'] += 1

                    # Support Story Calculation.............. but It takes so long time ....
                    if(storyCalSupport) :
                        try :
                            sql = "'Epic Link' = " + epic['Epic Key']
                            setfield = [ 'status']
                            storyissue = getFilterQueryResult(dev_jira, sql, getFieldList=setfield)
                        except KeyError as ke:
                            print("[Exception] JIRA Key Error : Initiative Key : {0}, Epic Key ==> {1}, Error Code = {2}".format(initiative['Initiative Key'], epic['Epic Key'], ke))
                            log.write("[Exception] JIRA Key Error Epic Key ==> {0}, Error Code = {1}".format(initiative['Initiative Key'], epic['Epic Key'], ke))
                            epic['STORY'] = 0
                            epic['StoryCnt'] = 0
                            epic['StoryResolutionCnt'] = 0
                            initiative['StoryCnt'] += epic['StoryCnt']
                            initiative['StoryResolutionCnt'] += epic['StoryResolutionCnt']
                            continue
                            '''
                            #{ TVPLAT-XXXX : [ epic1, epic2, ... ] , TVPLAT-XXXX : [ epic1, epic2, ... ] }
                            if initiative['Initiative Key'] in ExceptionList.keys() :
                                ExceptionList["{0}".format(initiative['Initiative Key'])].append(epic['Epic Key'])
                            else :
                                ExceptionList["{0}".format(initiative['Initiative Key'])] = [epic['Epic Key']]
                            print("ExeptionList = ", ExceptionList)
                            continue
                            '''
                        except JIRAError as e:
                            print("[Exception] JIRAError == Error Epic Key ==> ", epic['Epic Key'] + " sql = ", sql)
                            log.write("[Exception] JIRAError Epic Key ==> {0}, Error Code = {1}".format(epic['Epic Key'], e.status_code))
                        else :
                            epic['STORY'] = getStorynTaskKeyListfromJira(storyissue)
                            epic['StoryCnt'] = len(epic['STORY'])
                            epic['StoryResolutionCnt'] = getChildStorynTaskResolvedCntfromJira(storyissue)
                            initiative['StoryCnt'] += epic['StoryCnt']
                            initiative['StoryResolutionCnt'] += epic['StoryResolutionCnt']
                        finally :
                            #print("finally ok == Epic Key ==> ", epic['Epic Key'] + " sql = ", sql)
                            pass

                    if(initative_due < conversionSprnitToDatetime(epic['TVSP'][updateSP])) :
                        if(epic['Status'] != "Delivered" and epic['Status'] != "Resolved" and epic['Status'] != "Closed" and epic['Status'] != "Deferred") :
                            if(epic['GovOrDeployment'] != "O") :
                                initiative['AbnormalEpicSprint'] = "O"
                                epic['AbnormalEpicSprint'] = "O"


            #print(initiative)
            resultDB.append(initiative)
        else :
            print("Can't find Initiative or Epic Key from JIRA Filter Result.")
            pass

    return resultDB


#===========================================================================
# Get a initiative / Epic issue from Jira
# [param] filterResult : Jira Result from Filtered JIRA Query (Initiative or Epic Filter)
# [param] InitiativeKey : Initiative or Epic Key
# [return] Initiative jira issue
#===========================================================================
def getJiraInfo(filterResult, KeyID) :
    for dissue in filterResult :
        if(KeyID == getKey(dissue)) :
            return dissue
    return False



#===========================================================================
# Get Jira Result from FilterID
# [param] jiraHandle : Dev Jira handle
# [param] filterid : filter ID
# [return] JIRA Filtered Result
#===========================================================================
def getFilteredInitiativeInfofromJira(jiraHandle, filterid) :
    #Filter ID
    Initiative_webOS45_Initial_Dev = filterid

    #setfield = ('summary, duedate, assignee, status, created, components, labels')
    #result = getFilterIDResult(jiraHandle, Initiative_webOS45_Initial_Dev, setfield)
    if(Initiative_webOS45_Initial_Dev == 42101) :
        result = getFilterIDResult(jiraHandle, Initiative_webOS45_Initial_Dev)
    else :
        #setfield = ('summary, assignee, duedate, created, labels, status')
        setfield = ['summary', 'assignee', 'duedate', 'created', 'labels', 'status', 'issuelinks', 'resolutiondate', 'components' ]
        result = getFilterIDResult(jiraHandle, Initiative_webOS45_Initial_Dev, setfield)

    return result


#===========================================================================
# Python code to remove duplicate elements
# [param] duplicate : List with dulplicated Data
# [return] List Data
#===========================================================================
def RemoveDuplicateInList(duplicate):
    final_list = []
    for num in duplicate:
        if num not in final_list:
            final_list.append(num)
    return final_list


#===========================================================================
# Get Difference between List A and List B
# [param] listA : List Data
# [param] listB : List Data
# [return] List Data
#===========================================================================
def getDiffList(listA, listB) :
    return (list(set(listA) - set(listB)))




#===========================================================================
# Get the detail Initiative Information needed for history management from Jira
# [param] finalinfo : Initative Detail Info after updateing data from excel.
#   final_info = [ default_initiative_info1, default_initiative_info2, default_initiative_info3, ..., default_initiative_infoN ]
#   default_initiative_info = {
#        'Initiative Key' : '',
#        'Summary' : '',
#        'Assignee' : '',
#        'Status' : '',
#        'Release_SP' : '',
#        'CreatedDate' : '',
#        '관리대상' : '',
#        'Risk 관리 대상' : '',
#        'Initiative Order' : '',
#        'EPIC' : [],
#        'DEMO' : [],
#        'CCC' : [],
#        'TestCase' : [],
#        'Dev_Verification' : [],
#        'TVSP' : {},
#        }
#
#    default_epic_info = {
#            'Epic Key' : '',
#            'Release_SP' : '',
#            'Summary' : "",
#            'Assignee' : '',
#            'duedate' : '',
#            'Status' : '',
#            'CreatedDate' : '',
#            'TVSP' : { },
#        }
#
# [return] Final Data : all initiative_info list []
#===========================================================================
def updateInitiativeDetailInfoToXls(Sheetname, finalinfo) :
    row_index = 3
    col_index = 1
    index = 1

    for initiative in finalinfo :
        initlinkinfo = "http://hlm.lge.com/issue/browse/"+str(initiative['Initiative Key'])
        # write No Column
        setXlsCell(Sheetname, row_index, CI_No, index, False, None)
        # write 관리대상 Column
        setXlsCell(Sheetname, row_index, CI_SPE_M, initiative['관리대상'], False, None)
        # write 관리대상 Column
        setXlsCell(Sheetname, row_index, CI_SPE_R, initiative['Risk 관리 대상'], False, None)

        # write Initiative Order
        setXlsCell(Sheetname, row_index, CI_InitOrder, initiative['Initiative Order'], False, None)
        # write Issue Type
        setXlsCell(Sheetname, row_index, CI_IssueType, 'Initiative', False, None)
        # write Initiative Key
        setXlsCell(Sheetname, row_index, CI_InitKey, initiative['Initiative Key'], False, initlinkinfo)
        # write Epic Key
        setXlsCell(Sheetname, row_index, CI_EpicKey, initiative['Initiative Key'], False, initlinkinfo)
        # write Summary
        setXlsCell(Sheetname, row_index, CI_Summary, initiative['Summary'], False, None)

        # write Owner
        func = '=VLOOKUP(J{0},조직정보!$A$1:$E$5000,2,FALSE)'.format(row_index)
        setXlsCell(Sheetname, row_index, CI_Owner, func, False, None)
        # write Assignee
        setXlsCell(Sheetname, row_index, CI_Assignee, initiative['Assignee'], False, None)
        # write 조직책임자
        func = '=VLOOKUP(J{0},조직정보!$A$1:$E$5000,5,FALSE)'.format(row_index)
        setXlsCell(Sheetname, row_index, CI_OrganizationLeader, func, False, None)
        # write 조직
        func = '=VLOOKUP(J{0},조직정보!$A$1:$E$5000,4,FALSE)'.format(row_index)
        setXlsCell(Sheetname, row_index, CI_Organization, func, False, None)

        # write Status
        setXlsCell(Sheetname, row_index, CI_Status, initiative['Status'], False, None)
        # write CreatedDate
        setXlsCell(Sheetname, row_index, CI_Created, initiative['CreatedDate'], False, None)
        # write Release Sprint
        setXlsCell(Sheetname, row_index, CI_ReleaseSP, initiative['Release_SP'], False, None)
        # write 비고
        func = '=IF(O{0}=P{1}, "", "일정변경")'.format(row_index, row_index)
        setXlsCell(Sheetname, row_index, CI_CM, func, False, None)
        # write 금일기준
        setXlsCell(Sheetname, row_index, CI_Today, initiative['TVSP'][updateSP], False, None)

        # write SP
        initiative['RescheduleCnt'] = 0
        schedule = initiative['Release_SP']
        for key, value in initiative['TVSP'].items() :
            colpos = getColumnIndex(Sheetname, 2, key)
            setXlsCell(Sheetname, row_index, colpos, value, False, None)
            if(schedule != value and (value != "미설정" and value != "기완료" and value != "" and value != "TVSP_UNDEF" and value != "None")) :
                #print("++ Case : Initiative key = {}, schedule = {}, value = {}".format(initiative['Initiative Key'], schedule, value))
                initiative['RescheduleCnt'] += 1
                schedule =  value

        # write 일정변경Cnt
        setXlsCell(Sheetname, row_index, CI_ReschedCnt, initiative['RescheduleCnt'], False, None)


        # write EpicCnt
        setXlsCell(Sheetname, row_index, CI_EpicCnt, initiative['EpicCnt'], False, None)
        # write EpicResolutionCnt
        setXlsCell(Sheetname, row_index, CI_EpicResolutionCnt, initiative['EpicResolutionCnt'], False, None)
        # write DelayedCnt
        setXlsCell(Sheetname, row_index, CI_EpicDelayedCnt, initiative['EpicDelayedCnt'], False, None)
        # write EpicResolutionRate
        if(initiative['EpicCnt'] > 0) :
            EpicRate = int(initiative['EpicResolutionCnt']*100 / initiative['EpicCnt'])
        else :
            EpicRate = 0
        EpicRate = "{0}%".format(EpicRate)
        setXlsCell(Sheetname, row_index, CI_EpicResolutionRate, EpicRate, False, None)
        # write StoryCnt
        setXlsCell(Sheetname, row_index, CI_StoryCnt, initiative['StoryCnt'], False, None)
        # write StoryResolutionCnt
        setXlsCell(Sheetname, row_index, CI_StoryReolutionCnt, initiative['StoryResolutionCnt'], False, None)
        # write StoryResolutionRate
        if(initiative['StoryCnt'] > 0) :
            StoryRate = int(initiative['StoryResolutionCnt']*100 / initiative['StoryCnt'])
        else :
            StoryRate = 0
        StoryRate = "{0}%".format(StoryRate)
        setXlsCell(Sheetname, row_index, CI_StoryReolutionRate, StoryRate, False, None)

        # write RMS
        func = '=IF(ISNUMBER(FIND("RMS",H{0})), "O","")'.format(row_index)
        setXlsCell(Sheetname, row_index, CI_RMS, func, False, None)

        # write STE On Site
        setXlsCell(Sheetname, row_index, CI_STEOnSite, initiative['STEOnSite'], False, None)

        # write Scope of Change
        setXlsCell(Sheetname, row_index, CI_Scope, initiative['ScopeOfChange'], False, None)

        # write Abnormal Epic
        setXlsCell(Sheetname, row_index, CI_AbEpicSprint, initiative['AbnormalEpicSprint'], False, None)

        # write Governing or Deployment Issue
        setXlsCell(Sheetname, row_index, CI_GovOrDeployment, initiative['GovOrDeployment'], False, None)

        row_index += 1
        index += 1
        for epic in initiative['EPIC'] :
            StoryRate = 0
            initlinkinfo = "http://hlm.lge.com/issue/browse/"+str(initiative['Initiative Key'])
            epiclinkinfo = "http://hlm.lge.com/issue/browse/"+str(epic['Epic Key'])
            # write No Column
            setXlsCell(Sheetname, row_index, CI_No, index, False, None)

            # write Initiative Order
            setXlsCell(Sheetname, row_index, CI_InitOrder, initiative['Initiative Order'], False, None)
            # write Issue Type
            setXlsCell(Sheetname, row_index, CI_IssueType, 'EPIC', False, None)
            # write Initiative Key
            setXlsCell(Sheetname, row_index, CI_InitKey, initiative['Initiative Key'], False, initlinkinfo)
            # write Epic Key
            setXlsCell(Sheetname, row_index, CI_EpicKey, epic['Epic Key'], False, epiclinkinfo)
            # write Summary
            setXlsCell(Sheetname, row_index, CI_Summary, epic['Summary'], False, None)

            # write Owner
            func = '=VLOOKUP(J{0},조직정보!$A$1:$E$5000,2,FALSE)'.format(row_index)
            setXlsCell(Sheetname, row_index, CI_Owner, func, False, None)
            # write Assignee
            setXlsCell(Sheetname, row_index, CI_Assignee, epic['Assignee'], False, None)
            # write 조직책임자
            func = '=VLOOKUP(J{0},조직정보!$A$1:$E$5000,5,FALSE)'.format(row_index)
            setXlsCell(Sheetname, row_index, CI_OrganizationLeader, func, False, None)
            # write 조직
            func = '=VLOOKUP(J{0},조직정보!$A$1:$E$5000,4,FALSE)'.format(row_index)
            setXlsCell(Sheetname, row_index, CI_Organization, func, False, None)

            # write Status
            setXlsCell(Sheetname, row_index, CI_Status, epic['Status'], False, None)
            # write CreatedDate
            setXlsCell(Sheetname, row_index, CI_Created, epic['CreatedDate'], False, None)
            # write Release Sprint
            setXlsCell(Sheetname, row_index, CI_ReleaseSP, epic['Release_SP'], False, None)
            # write StoryCnt
            setXlsCell(Sheetname, row_index, CI_StoryCnt, epic['StoryCnt'], False, None)
            # write StoryResolutionCnt
            setXlsCell(Sheetname, row_index, CI_StoryReolutionCnt, epic['StoryResolutionCnt'], False, None)

            # write 비고
            func = '=IF(O{0}=P{1}, "", "일정변경")'.format(row_index, row_index)
            setXlsCell(Sheetname, row_index, CI_CM, func, False, None)

            # write 금일기준
            # 20180516
            #print("금일기준 : epic key = {}, updateSP = {}, epic['TVSP'] = {}".format(epic['Epic Key'], updateSP, epic['TVSP']))
            setXlsCell(Sheetname, row_index, CI_Today, epic['TVSP'][updateSP], False, None)

            # write RMS
            func = '=IF(ISNUMBER(FIND("RMS",H{0})), "O","")'.format(row_index)
            setXlsCell(Sheetname, row_index, CI_RMS, func, False, None)

            # write StoryResolutionRate
            if(epic['StoryCnt'] > 0) :
                StoryRate = int(epic['StoryResolutionCnt']*100 / epic['StoryCnt'])
            else :
                StoryRate = 0
            StoryRate = "{0}%".format(StoryRate)
            setXlsCell(Sheetname, row_index, CI_StoryReolutionRate, StoryRate, False, None)

            # write SP
            epic['RescheduleCnt'] = 0
            schedule = epic['Release_SP']
            for key, value in epic['TVSP'].items() :
                colpos = getColumnIndex(Sheetname, 2, key)
                setXlsCell(Sheetname, row_index, colpos, value, False, None)
                if(schedule != value and (value != "미설정" and value != "기완료" and value != "" and value != "TVSP_UNDEF" and value != "None")) :
                    #print("++ Case : epic key = {}, schedule = {}, value = {}".format(epic['Epic Key'], schedule, value))
                    epic['RescheduleCnt'] += 1
                    schedule =  value

            # write 일정변경Cnt
            setXlsCell(Sheetname, row_index, CI_ReschedCnt, epic['RescheduleCnt'], False, None)
            # write Abnormal Epic
            setXlsCell(Sheetname, row_index, CI_AbEpicSprint, epic['AbnormalEpicSprint'], False, None)
            # write Governing or Deployment Issue
            setXlsCell(Sheetname, row_index, CI_GovOrDeployment, epic['GovOrDeployment'], False, None)

            row_index += 1
            index += 1



#===========================================================================
# Clear data within region
# [param] Sheetname : Excel Sheet Name
# [param] start_row : Start Row Position of region to be cleared
# [param] start_col : Start Column Position of region to be cleared
# [return] None
#===========================================================================
def prepareNewXlsSheet(Sheetname, start_row, start_col) :
    for x in range(start_row, maxResultCnt) :
        for y in range(start_col, 50) :
            setXlsCell(Sheetname, x, y, '', False, None)



#===========================================================================
# Main Function
# [param] None
# [return] None
#===========================================================================
if __name__ == "__main__" :
    # jira Handle open
    dev_jira = JIRA(DevTracker, basic_auth = (ID, PASSWORD))

    # create log file
    if (os.path.isfile(logfilename)) :
        os.remove(logfilename)

    log = open(logfilename, 'wt', encoding='utf8')

    # Create Excel workbook
    workbook = xlsrd.load_workbook(openfilename)
    org_sheet = workbook["최종"]

    workbook.copy_worksheet(org_sheet)
    cur_sheet = workbook["최종 Copy"]
    #print(workbook.get_sheet_names())
    cur_sheet.title = "금일작업본"

    # set max row/column count
    MAX_RowCount = getRowCount(cur_sheet, 3, 1)
    MAX_ColCount = getColumnCount(cur_sheet, 2, 1)
    log.write("# Max_RowCount = {0}, MAX_ColCount = {1}".format(MAX_RowCount, MAX_ColCount))

    # set title column index to variables
    CI_No = getColumnIndex(cur_sheet, 2, "No")
    CI_SPE_M = getColumnIndex(cur_sheet, 2, "관리대상")
    CI_SPE_R = getColumnIndex(cur_sheet, 2, "Risk 관리 대상")
    CI_InitOrder = getColumnIndex(cur_sheet, 2, "Initiative Order")
    CI_IssueType = getColumnIndex(cur_sheet, 2, "Type")
    CI_InitKey = getColumnIndex(cur_sheet, 2, "Initiative Key")
    CI_EpicKey = getColumnIndex(cur_sheet, 2, "Epic Key")
    CI_Summary = getColumnIndex(cur_sheet, 2, "Summary")
    CI_Owner = getColumnIndex(cur_sheet, 2, "Owner")
    CI_Assignee = getColumnIndex(cur_sheet, 2, "Assignee")
    CI_OrganizationLeader = getColumnIndex(cur_sheet, 2, "조직책임자")
    CI_Organization = getColumnIndex(cur_sheet, 2, "조직")
    CI_Status = getColumnIndex(cur_sheet, 2, "Status")
    CI_Today = getColumnIndex(cur_sheet, 2, "금일기준")
    CI_Created = getColumnIndex(cur_sheet, 2, "CreatedDate")
    CI_ReleaseSP = getColumnIndex(cur_sheet, 2, "Release_SP")
    CI_CM = getColumnIndex(cur_sheet, 2, "비고")
    CI_StartPos = getColumnIndex(cur_sheet, 2, startSP)
    CI_EndPos = getColumnIndex(cur_sheet, 2, endSP)

    CI_EpicCnt = getColumnIndex(cur_sheet, 2, "EpicCnt")
    CI_EpicResolutionCnt = getColumnIndex(cur_sheet, 2, "EpicResolutionCnt")
    CI_EpicResolutionRate = getColumnIndex(cur_sheet, 2, "Epic진행율")
    CI_EpicDelayedCnt = getColumnIndex(cur_sheet, 2, "DelayedEpicCnt")
    CI_StoryCnt = getColumnIndex(cur_sheet, 2, "StoryCnt")
    CI_StoryReolutionCnt = getColumnIndex(cur_sheet, 2, "StoryReolutionCnt")
    CI_StoryReolutionRate = getColumnIndex(cur_sheet, 2, "Story진행율")
    CI_RMS = getColumnIndex(cur_sheet, 2, "RMS")
    CI_Scope = getColumnIndex(cur_sheet, 2, "Scope")
    CI_ReschedCnt = getColumnIndex(cur_sheet, 2, "ReschedCnt")
    CI_STEOnSite = getColumnIndex(cur_sheet, 2, "STE투입")
    CI_AbEpicSprint = getColumnIndex(cur_sheet, 2, "Abnormal Epic Sprint")
    CI_GovOrDeployment = getColumnIndex(cur_sheet, 2, "GovOrDeployment")


    #print("Type = {0}, Epic Key = {1}, Initiative Key = {2}, Release_SP = {3}".format(CI_IssueType, CI_EpicKey, CI_InitKey, CI_ReleaseSP))

    TVSP11_Start = conversionCreatedDateToDatetime("2018-1-15")
    TVSP12_Start = conversionCreatedDateToDatetime("2018-1-29")
    TVSP13_Start = conversionCreatedDateToDatetime("2018-2-12")
    TVSP14_Start = conversionCreatedDateToDatetime("2018-2-26")
    TVSP15_Start = conversionCreatedDateToDatetime("2018-3-12")
    TVSP16_Start = conversionCreatedDateToDatetime("2018-4-2")
    TVSP17_Start = conversionCreatedDateToDatetime("2018-4-16")
    TVSP18_Start = conversionCreatedDateToDatetime("2018-4-30")
    TVSP19_Start = conversionCreatedDateToDatetime("2018-5-14")
    TVSP20_Start = conversionCreatedDateToDatetime("2018-5-28")
    TVSP21_Start = conversionCreatedDateToDatetime("2018-6-11")
    TVSP22_Start = conversionCreatedDateToDatetime("2018-6-25")
    TVSP23_Start = conversionCreatedDateToDatetime("2018-7-9")
    TVSP24_Start = conversionCreatedDateToDatetime("2018-7-23")
    TVSP25_Start = conversionCreatedDateToDatetime("2018-8-6")
    TVSP26_Start = conversionCreatedDateToDatetime("2018-8-20")
    TVSP27_Start = conversionCreatedDateToDatetime("2018-9-3")
    TVSP28_Start = conversionCreatedDateToDatetime("2018-9-17")
    TVSP29_Start = conversionCreatedDateToDatetime("2018-10-1")
    TVSP30_Start = conversionCreatedDateToDatetime("2018-10-15")
    TVSP31_Start = conversionCreatedDateToDatetime("2018-10-29")
    TVSP32_Start = conversionCreatedDateToDatetime("2018-11-12")

    #getInitiativeRowIndex(cur_sheet, 'TVPLAT-874')


    # Initiative ========================================================================
    # 1. JIRA에서 Initiative Filter에 맞는 Initiative Key를 Jira로 구성한다. [ 'key1', 'key2', .... ]
    log.write("\n\n# 1. JIRA에서 Initiative Filter에 맞는 Initiative Key를 Jira로 구성한다. [ 'key1', 'key2', .... ]\n")
    print("\n\n# 1. JIRA에서 Initiative Filter에 맞는 Initiative Key를 Jira로 구성한다. [ 'key1', 'key2', .... ]\n")
    Initiative_FilterResult = getFilteredInitiativeInfofromJira(dev_jira, 42101)
    jira_initiative_keylist = getInitiativeKeylistFromJira(Initiative_FilterResult)

    # 2. Excel로 부터 Initiative Key List를 구성한다. [ 'key1', 'key2', .... ]
    log.write("\n\n# 2. Excel({0})로 부터 Initiative Key List를 구성한다. [ 'key1', 'key2', .... ]\n".format(openfilename))
    print("\n\n# 2. Excel({0})로 부터 Initiative Key List를 구성한다. [ 'key1', 'key2', .... ]\n".format(openfilename))
    xls_initiative_keylist = getInitiativeKeylistFromXls(cur_sheet, 3)

    # 3. Jira상의 Initiative Key List와 엑셀상에 관리되는 Initiative Key List를 비교한다.
    log.write("\n\n# 3. Jira상의 Initiative Key List와 엑셀상에 관리되는 Initiative Key List를 비교한다.\n")
    log.write("\n################## New Initiative List (JIRA - Excel) ##################\n")
    print("\n################## New Initiative List (JIRA - Excel) ##################")
    newkey = getDiffList(jira_initiative_keylist, xls_initiative_keylist)
    print(newkey)
    log.write(str(newkey))
    print("\n################## Del Initiative List (Excel - JIRA) ##################")
    log.write("\n################## Del Initiative List (Excel - JIRA) ##################\n")
    delkey = getDiffList(xls_initiative_keylist, jira_initiative_keylist)
    print(delkey)
    log.write(str(delkey))


    # Epic ==============================================================================
    # 4. JIRA에서 Initative IssueLinks 정보에서 Epic Key List를 구성한다. [ 'key1', 'key2', .... ]
    log.write("\n\n# 4. JIRA에서 Initative IssueLinks 정보에서 Epic Key List를 구성한다. [ 'key1', 'key2', .... ]\n")
    jira_Issuelinks_epic_keylist = getEpicKeyListfromJira(Initiative_FilterResult, "Initiative_Filter")

    # 5. JIRA에서 Epic Filter에 맞는 Epic Key List를 구성한다. [ 'key1', 'key2', .... ]
    log.write("\n\n# 5. JIRA에서 Epic Filter에 맞는 Epic Key List를 구성한다. [ 'key1', 'key2', .... ]\n")
    Epic_FilterResult = getFilteredInitiativeInfofromJira(dev_jira, 42317)
    jira_epic_keylist = getEpicKeyListfromJira(Epic_FilterResult, "Epic_Filter")
    xls_epic_keylist = getEpicKeyListfromXls(cur_sheet, 3)
    log.write("\nEpic : Jira = {0} EA, Excel = {1} EA\n".format(len(jira_epic_keylist), len(xls_epic_keylist)))

    # 6. JIRA에서 Epic Filter를 이용해 구성한 정보와 Initiative Issuelinks[] 정보를 이용해 만든 Epic Key List가 일치하는지를 체크한다. [ 'key1', 'key2', .... ]
    log.write("\n\n# 6. JIRA에서 Epic Filter를 이용해 구성한 정보와 Initiative Issuelinks[] 정보를 이용해 만든 Epic Key List가 일치하는지를 체크한다. [ 'key1', 'key2', .... ]\n")
    log.write("\n################## Compare1 Epic List (jira filter - jira link) ##################\n")
    print("\n################## Compare1 Epic List (jira filter - jira link) ##################")
    new_issuelinks_epickey = getDiffList(jira_epic_keylist, jira_Issuelinks_epic_keylist)
    print(new_issuelinks_epickey)
    log.write(str(new_issuelinks_epickey))
    log.write("\n################## Compare2 Epic List (jira link - jira filter) ##################\n")
    print("\n################## Compare2 Epic List (jira link - jira filter) ##################")
    new_filtered_epickey = getDiffList(jira_Issuelinks_epic_keylist, jira_epic_keylist)
    print(new_filtered_epickey)
    log.write(str(new_filtered_epickey))

    log.write("\n################## New Epic List (JIRA - Excel) ##################\n")
    print("\n################## New Epic List (JIRA - Excel) ##################")
    newEpickey = getDiffList(jira_epic_keylist, xls_epic_keylist)
    print(newEpickey)
    log.write(str(newEpickey))
    log.write("\n################## Del Epic List (Excel - JIRA) ##################\n")
    print("\n################## Del Epic List (Excel - JIRA) ##################")
    delEpickey = getDiffList(xls_epic_keylist, jira_epic_keylist)
    log.write(str(delEpickey))
    print(delEpickey)

    # 7. 각 Initiative 하위에 존재하는 Epick List 구성
    log.write("\n\n# 7. 각 Initiative 하위에 존재하는 Epick List 구성\n")
    # list[ { 'key' : 'Initative Key',  'epiclist' : [ 'Epic Key1', 'Epic Key2', ... ]}, ....  ]
    print("\n################## Prepare Initative - Epic List from Jira ##################")
    jira_Init_EpicLists = getInitiativeAllEpicsListfromJira(Initiative_FilterResult)
    #print(jira_Init_EpicLists)
    log.write(str(jira_Init_EpicLists))


    # 8. Jira로 부터 얻은 Initiative Key를 가지고 엑셀의 정보를 먼저 Update 한다. (Jira 기준 - Initative Key List[])
    #    Update Detail Initative(Epic) Information from Excel first.
    #    (Release SP, SP History, 관리대상, 관리대상 Risk, ?Epic List?)
    log.write("\n\n# 8. Jira로 부터 얻은 Initiative Key를 가지고 엑셀의 정보를 먼저 Update 한다. (Jira 기준 - Initative Key List[])\n")
    print("\n################## Prepare Initative DB : Update History Data from Excel ##################")
    tmp_Initiative = getInitiativeDetailInfofromXls(cur_sheet, jira_initiative_keylist, jira_Init_EpicLists)
    #print(tmp_Initiative)
    log.write(str(tmp_Initiative))


    # 9. Initiative Key를 기준으로 Jira상의 최신 정보를 Update한다.
    log.write("\n\n# 9. Initiative Key를 기준으로 Jira상의 최신 정보를 Update한다.\n")
    print("\n################## Update Initative DB : Update Latest Data from Jira ##################\n")
    finalDB = getInitiativeDetailInfofromJira(Initiative_FilterResult, Epic_FilterResult, tmp_Initiative)
    #print(finalDB)
    log.write(str(finalDB))

    #9-1. Story 가져오다 Error 발생한 Initiative / Epic 은 다시 1번 더 Story 계산하는 work arround
    print("################# Exception List ##########################")
    print(ExceptionList)
    '''
    for k, v in ExceptionList.items() :
        for initdict in finalDB :
            initdict[]
            epic['STORY'] = getStorynTaskKeyListfromJira(storyissue)
            epic['StoryCnt'] = len(epic['STORY'])
            epic['StoryResolutionCnt'] = getChildStorynTaskResolvedCntfromJira(storyissue)
            initiative['StoryCnt'] += epic['StoryCnt']
            initiative['StoryResolutionCnt'] += epic['StoryResolutionCnt']
    '''

    # 10. Jira로 부터 얻은 Initiative Key를 기준으로 Jira상의 최신 정보를 Excel 문서에 Update 한다.
    log.write("\n\n# 10. Jira로 부터 얻은 Initiative Key를 기준으로 Jira상의 최신 정보를 Excel 문서에 Update 한다.\n")
    print("\n################## Make New Sheet and Clear Excel Area ##################")
    prepareNewXlsSheet(cur_sheet, 3, 1)
    print("\n################## Write Final DB Info to Excel Sheet ##################")
    updateInitiativeDetailInfoToXls(cur_sheet, finalDB)

    # 11. Jira상의 최신 정보를 Excel 문서에 Update 후 별도 Excel File로 저장한다.
    log.write("\n\n# 11. Jira상의 최신 정보를 Excel 문서에 Update 후 별도 Excel File로 저장한다.\n")
    print("\n################## Save Excel Sheet ##################")
    cur_sheet.auto_filter.ref = 'A2:AZ2'
    workbook.save(savefilename)
    log.close()
    os.system('start excel.exe %s' % savefilename)
